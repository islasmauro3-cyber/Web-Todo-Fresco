import os
import re
import csv
import io
import time
import uuid
import hashlib
import json
from datetime import datetime, date
from functools import wraps

from dotenv import load_dotenv
load_dotenv()

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from flask import (
    Flask, render_template, request, redirect, url_for,
    session, flash, send_from_directory, abort, Response
)
from werkzeug.utils import secure_filename
from PIL import Image

from database import get_db, init_db

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "cambia-esta-clave-secreta-123")

UPLOAD_FOLDER = os.path.join(app.root_path, "static", "uploads")
ALLOWED_EXTENSIONS = {"jpg", "jpeg", "png", "webp", "gif"}
MAX_IMAGE_SIZE = (800, 800)

APP_PASSWORD = os.environ.get("APP_PASSWORD", "admin123")

# ── Twilio config ──────────────────────────────────────────────────────────────
TWILIO_ACCOUNT_SID  = os.environ.get("TWILIO_ACCOUNT_SID", "")
TWILIO_AUTH_TOKEN   = os.environ.get("TWILIO_AUTH_TOKEN", "")
TWILIO_WA_FROM      = os.environ.get("TWILIO_WA_FROM", "")      # whatsapp:+17372508034
TWILIO_WA_OWNER     = os.environ.get("TWILIO_WA_OWNER", "")     # owner's personal number
# SID de la plantilla Twilio para mensajes iniciados (sandbox: "Marketing Promotions")
TWILIO_CONTENT_SID  = os.environ.get("TWILIO_CONTENT_SID", "HXb5b62575e6e4ff6129ad7c8efe01a50f")

print(f"[Twilio] CONTENT_SID cargado: {TWILIO_CONTENT_SID!r}", flush=True)

UNIDADES = [
    "por kg",
    "por unidad",
    "por 100g",
    "por 1/4 kg",
    "por media horma",
    "por horma entera",
]

_UNIDAD_MAP = {
    "kg": "por kg", "kilo": "por kg", "kilos": "por kg", "kilogramo": "por kg",
    "por kg": "por kg",
    "100g": "por 100g", "100 g": "por 100g", "por 100g": "por 100g",
    "1/4": "por 1/4 kg", "cuarto": "por 1/4 kg", "1/4 kg": "por 1/4 kg",
    "por 1/4 kg": "por 1/4 kg", "cuarto kg": "por 1/4 kg",
    "media horma": "por media horma", "1/2 horma": "por media horma",
    "por media horma": "por media horma", "media": "por media horma",
    "horma": "por horma entera", "horma entera": "por horma entera",
    "por horma entera": "por horma entera",
    "unidad": "por unidad", "u": "por unidad", "ud": "por unidad",
    "por unidad": "por unidad",
}


@app.template_filter("unidad_corta")
def unidad_corta(u):
    """'por kg' → 'kg', para mostrar '$3500 / kg'."""
    return (u or "unidad").replace("por ", "", 1)


def allowed_file(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS


def save_image(file):
    ext = file.filename.rsplit(".", 1)[1].lower()
    filename = f"{uuid.uuid4().hex}.{ext}"
    filepath = os.path.join(UPLOAD_FOLDER, filename)
    img = Image.open(file)
    img.thumbnail(MAX_IMAGE_SIZE, Image.LANCZOS)
    img.save(filepath, optimize=True, quality=85)
    return filename


def delete_image(filename):
    if filename:
        path = os.path.join(UPLOAD_FOLDER, filename)
        if os.path.exists(path):
            os.remove(path)


def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get("autenticado"):
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return decorated


def oferta_activa(producto, variantes=None):
    """Determina si un producto tiene oferta activa hoy."""
    hoy = date.today().isoformat()
    desde = (producto.get("oferta_desde") or "")
    hasta = (producto.get("oferta_hasta") or "")
    if desde and hoy < desde:
        return False
    if hasta and hoy > hasta:
        return False
    if variantes is not None:
        return any(v["en_oferta"] for v in variantes)
    return bool(producto.get("en_oferta"))


# ── Variantes helpers ──────────────────────────────────────────────────────────

def _get_variantes(producto_id, db=None):
    close = db is None
    if close:
        db = get_db()
    rows = db.execute(
        "SELECT * FROM variantes WHERE producto_id=? ORDER BY orden, id", (producto_id,)
    ).fetchall()
    if close:
        db.close()
    return [dict(r) for r in rows]


def _get_variantes_bulk(producto_ids, db=None):
    close = db is None
    if close:
        db = get_db()
    if not producto_ids:
        if close:
            db.close()
        return {}
    placeholders = ",".join("?" * len(producto_ids))
    rows = db.execute(
        f"SELECT * FROM variantes WHERE producto_id IN ({placeholders}) ORDER BY producto_id, orden, id",
        producto_ids,
    ).fetchall()
    if close:
        db.close()
    result = {}
    for r in rows:
        r = dict(r)
        result.setdefault(r["producto_id"], []).append(r)
    return result


def _parse_variantes_form():
    """Extrae variantes del formulario (campos variante_precio_N / variante_unidad_N)."""
    indices = set()
    for key in request.form:
        m = re.match(r"^variante_precio_(\d+)$", key)
        if m:
            indices.add(int(m.group(1)))
    variantes = []
    for idx in sorted(indices):
        precio_raw = request.form.get(f"variante_precio_{idx}", "").strip().replace(",", ".")
        unidad = request.form.get(f"variante_unidad_{idx}", "por unidad")
        if unidad not in UNIDADES:
            unidad = "por unidad"
        en_oferta = 1 if request.form.get(f"variante_oferta_{idx}") else 0
        try:
            precio = float(precio_raw)
            if precio >= 0:
                variantes.append({
                    "unidad": unidad,
                    "precio": precio,
                    "en_oferta": en_oferta,
                    "orden": idx,
                })
        except ValueError:
            pass
    return variantes


def _guardar_variantes_db(producto_id, variantes, db):
    db.execute("DELETE FROM variantes WHERE producto_id=?", (producto_id,))
    for i, v in enumerate(variantes):
        db.execute(
            "INSERT INTO variantes (producto_id, unidad, precio, en_oferta, orden) VALUES (?,?,?,?,?)",
            (producto_id, v["unidad"], v["precio"], v["en_oferta"], i),
        )


# ── Autenticación ─────────────────────────────────────────────────────────────

@app.route("/login", methods=["GET", "POST"])
def login():
    if session.get("autenticado"):
        return redirect(url_for("dashboard"))
    error = None
    if request.method == "POST":
        if request.form.get("password") == APP_PASSWORD:
            session["autenticado"] = True
            session.permanent = True
            return redirect(url_for("dashboard"))
        error = "Contraseña incorrecta"
    return render_template("login.html", error=error)


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


# ── Dashboard ─────────────────────────────────────────────────────────────────

@app.route("/")
@login_required
def dashboard():
    db = get_db()
    total_productos = db.execute("SELECT COUNT(*) FROM productos WHERE activo=1").fetchone()[0]
    en_oferta = db.execute("""
        SELECT COUNT(DISTINCT p.id) FROM productos p
        JOIN variantes v ON v.producto_id = p.id
        WHERE p.activo=1 AND v.en_oferta=1
    """).fetchone()[0]
    total_clientes = db.execute("SELECT COUNT(*) FROM clientes WHERE activo=1").fetchone()[0]
    db.close()
    return render_template(
        "dashboard.html",
        total_productos=total_productos,
        en_oferta=en_oferta,
        total_clientes=total_clientes,
    )


# ── Productos ─────────────────────────────────────────────────────────────────

@app.route("/productos")
@login_required
def productos():
    db = get_db()
    rows = db.execute(
        "SELECT * FROM productos ORDER BY activo DESC, nombre ASC"
    ).fetchall()
    ids = [r["id"] for r in rows]
    variantes_by_prod = _get_variantes_bulk(ids, db)
    db.close()
    items = []
    for r in rows:
        p = dict(r)
        p["variantes"] = variantes_by_prod.get(p["id"], [])
        p["oferta_hoy"] = oferta_activa(p, p["variantes"])
        items.append(p)
    return render_template("productos.html", productos=items)


@app.route("/productos/nuevo", methods=["GET", "POST"])
@login_required
def producto_nuevo():
    if request.method == "POST":
        return _guardar_producto(None)
    return render_template("producto_form.html", producto=None, unidades=UNIDADES, variantes_form=[])


@app.route("/productos/<int:pid>/editar", methods=["GET", "POST"])
@login_required
def producto_editar(pid):
    db = get_db()
    row = db.execute("SELECT * FROM productos WHERE id=?", (pid,)).fetchone()
    variantes = _get_variantes(pid, db)
    db.close()
    if not row:
        abort(404)
    if request.method == "POST":
        return _guardar_producto(dict(row))
    return render_template(
        "producto_form.html", producto=dict(row), unidades=UNIDADES, variantes_form=variantes
    )


def _guardar_producto(existente):
    pid = existente["id"] if existente else None
    nombre = request.form.get("nombre", "").strip()
    descripcion = request.form.get("descripcion", "").strip()
    oferta_desde = request.form.get("oferta_desde") or None
    oferta_hasta = request.form.get("oferta_hasta") or None
    activo = 1 if request.form.get("activo") else 0

    variantes = _parse_variantes_form()

    def render_form(msg=None, category="danger"):
        if msg:
            flash(msg, category)
        return render_template(
            "producto_form.html", producto=existente,
            unidades=UNIDADES, variantes_form=variantes,
        )

    if not nombre:
        return render_form("El nombre es obligatorio.")
    if not variantes:
        return render_form("Debés agregar al menos un precio.")

    primera = variantes[0]
    precio = primera["precio"]
    unidad = primera["unidad"]
    en_oferta = 1 if any(v["en_oferta"] for v in variantes) else 0

    foto_actual = existente["foto"] if existente else None
    foto = foto_actual
    file = request.files.get("foto")
    if file and file.filename:
        if not allowed_file(file.filename):
            return render_form("Formato de imagen no permitido. Usá JPG, PNG o WEBP.")
        delete_image(foto_actual)
        foto = save_image(file)

    db = get_db()
    if pid:
        db.execute(
            """UPDATE productos SET nombre=?, descripcion=?, precio=?, foto=?,
               unidad=?, en_oferta=?, oferta_desde=?, oferta_hasta=?, activo=?
               WHERE id=?""",
            (nombre, descripcion, precio, foto, unidad,
             en_oferta, oferta_desde, oferta_hasta, activo, pid),
        )
        flash("Producto actualizado.", "success")
    else:
        cur = db.execute(
            """INSERT INTO productos (nombre, descripcion, precio, foto, unidad,
               en_oferta, oferta_desde, oferta_hasta, activo)
               VALUES (?,?,?,?,?,?,?,?,?)""",
            (nombre, descripcion, precio, foto, unidad,
             en_oferta, oferta_desde, oferta_hasta, activo),
        )
        pid = cur.lastrowid
        flash("Producto creado.", "success")

    _guardar_variantes_db(pid, variantes, db)
    db.commit()
    db.close()
    return redirect(url_for("productos"))


@app.route("/productos/<int:pid>/borrar", methods=["POST"])
@login_required
def producto_borrar(pid):
    db = get_db()
    row = db.execute("SELECT foto FROM productos WHERE id=?", (pid,)).fetchone()
    if row:
        delete_image(row["foto"])
        db.execute("DELETE FROM productos WHERE id=?", (pid,))
        db.commit()
        flash("Producto eliminado.", "success")
    db.close()
    return redirect(url_for("productos"))


@app.route("/productos/<int:pid>/toggle-oferta", methods=["POST"])
@login_required
def toggle_oferta(pid):
    db = get_db()
    variantes = db.execute(
        "SELECT id, en_oferta FROM variantes WHERE producto_id=?", (pid,)
    ).fetchall()
    if variantes:
        any_on = any(v["en_oferta"] for v in variantes)
        nuevo = 0 if any_on else 1
        db.execute("UPDATE variantes SET en_oferta=? WHERE producto_id=?", (nuevo, pid))
        db.execute("UPDATE productos SET en_oferta=? WHERE id=?", (nuevo, pid))
    else:
        row = db.execute("SELECT en_oferta FROM productos WHERE id=?", (pid,)).fetchone()
        if row:
            nuevo = 0 if row["en_oferta"] else 1
            db.execute("UPDATE productos SET en_oferta=? WHERE id=?", (nuevo, pid))
    db.commit()
    db.close()
    return redirect(url_for("productos"))


# ── Clientes ──────────────────────────────────────────────────────────────────

@app.route("/clientes")
@login_required
def clientes():
    db = get_db()
    rows = db.execute(
        "SELECT * FROM clientes ORDER BY activo DESC, nombre ASC"
    ).fetchall()
    db.close()
    return render_template("clientes.html", clientes=[dict(r) for r in rows])


@app.route("/clientes/nuevo", methods=["GET", "POST"])
@login_required
def cliente_nuevo():
    if request.method == "POST":
        return _guardar_cliente(None)
    return render_template("cliente_form.html", cliente=None)


@app.route("/clientes/<int:cid>/editar", methods=["GET", "POST"])
@login_required
def cliente_editar(cid):
    db = get_db()
    row = db.execute("SELECT * FROM clientes WHERE id=?", (cid,)).fetchone()
    db.close()
    if not row:
        abort(404)
    if request.method == "POST":
        return _guardar_cliente(dict(row))
    return render_template("cliente_form.html", cliente=dict(row))


def _guardar_cliente(existente):
    cid = existente["id"] if existente else None
    nombre = request.form.get("nombre", "").strip()
    celular = request.form.get("celular", "").strip()
    notas = request.form.get("notas", "").strip()
    activo = 1 if request.form.get("activo") else 0

    if not nombre or not celular:
        flash("Nombre y celular son obligatorios.", "danger")
        return render_template("cliente_form.html", cliente=existente)

    db = get_db()
    if cid:
        db.execute(
            "UPDATE clientes SET nombre=?, celular=?, notas=?, activo=? WHERE id=?",
            (nombre, celular, notas, activo, cid),
        )
        flash("Cliente actualizado.", "success")
    else:
        db.execute(
            "INSERT INTO clientes (nombre, celular, notas, activo) VALUES (?,?,?,?)",
            (nombre, celular, notas, activo),
        )
        flash("Cliente agregado.", "success")
    db.commit()
    db.close()
    return redirect(url_for("clientes"))


@app.route("/clientes/<int:cid>/borrar", methods=["POST"])
@login_required
def cliente_borrar(cid):
    db = get_db()
    db.execute("DELETE FROM clientes WHERE id=?", (cid,))
    db.commit()
    db.close()
    flash("Cliente eliminado.", "success")
    return redirect(url_for("clientes"))


@app.route("/clientes/<int:cid>/toggle", methods=["POST"])
@login_required
def cliente_toggle(cid):
    db = get_db()
    row = db.execute("SELECT activo FROM clientes WHERE id=?", (cid,)).fetchone()
    if row:
        db.execute("UPDATE clientes SET activo=? WHERE id=?", (1 - row["activo"], cid))
        db.commit()
    db.close()
    return redirect(url_for("clientes"))


# ── Vista previa WhatsApp ──────────────────────────────────────────────────────

@app.route("/whatsapp")
@login_required
def whatsapp_preview():
    db = get_db()
    productos_all = db.execute(
        "SELECT * FROM productos WHERE activo=1 ORDER BY nombre"
    ).fetchall()
    clientes_activos = db.execute(
        "SELECT * FROM clientes WHERE activo=1 ORDER BY nombre"
    ).fetchall()
    ids = [p["id"] for p in productos_all]
    variantes_by_prod = _get_variantes_bulk(ids, db)
    db.close()

    ofertas_hoy = []
    for p in productos_all:
        p = dict(p)
        variantes = variantes_by_prod.get(p["id"], [])
        p["variantes"] = variantes
        if oferta_activa(p, variantes):
            ofertas_hoy.append(p)

    mensaje_base = _generar_mensaje(ofertas_hoy)
    mensaje_tpl = f"Hola {{{{nombre}}}}! 👋\n\n{mensaje_base}"
    total_clientes = len(clientes_activos)
    return render_template(
        "whatsapp_preview.html",
        ofertas=ofertas_hoy,
        clientes=clientes_activos,
        mensaje=mensaje_base,
        mensaje_tpl=mensaje_tpl,
        twilio_ok=_twilio_configurado(),
        twilio_owner=TWILIO_WA_OWNER,
        total_clientes=total_clientes,
        content_sid=TWILIO_CONTENT_SID,
    )


def _generar_mensaje(ofertas):
    if not ofertas:
        return "No hay productos en oferta hoy."
    lineas = ["🛍️ *¡Ofertas del día!* 🛍️", ""]
    for p in ofertas:
        lineas.append(f"✅ *{p['nombre']}*")
        if p["descripcion"]:
            lineas.append(f"   {p['descripcion']}")
        variantes = p.get("variantes") or []
        if variantes:
            for v in variantes:
                precio_fmt = f"${v['precio']:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
                u = (v.get("unidad") or "por unidad").replace("por ", "", 1)
                mark = " 🔥" if v.get("en_oferta") else ""
                lineas.append(f"   💲 {precio_fmt} / {u}{mark}")
        else:
            precio_fmt = f"${p['precio']:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
            u = (p.get("unidad") or "por unidad").replace("por ", "", 1)
            lineas.append(f"   💲 {precio_fmt} / {u}")
        lineas.append("")
    lineas.append("¡Contactame para hacer tu pedido! 😊")
    return "\n".join(lineas)


# ── Twilio / WhatsApp envío ───────────────────────────────────────────────────

def _twilio_configurado():
    return bool(TWILIO_ACCOUNT_SID and TWILIO_AUTH_TOKEN and TWILIO_WA_FROM)


def _get_twilio_client():
    if not _twilio_configurado():
        return None
    from twilio.rest import Client
    return Client(TWILIO_ACCOUNT_SID, TWILIO_AUTH_TOKEN)


def _es_error_contentsid(err):
    e = err.lower()
    return "contentsid" in e or "content_sid" in e or "63016" in e or "mbs" in e


def _enviar_via_template(client, to_wa, mensaje, content_sid):
    try:
        msg = client.messages.create(
            from_=TWILIO_WA_FROM,
            to=to_wa,
            content_sid=content_sid,
            content_variables=json.dumps({"1": mensaje}, ensure_ascii=False),
        )
        return msg.sid, None
    except Exception as e:
        return None, str(e)


def _enviar_whatsapp(celular, mensaje, content_sid=None):
    """Envía un mensaje WA. Devuelve (sid, None) o (None, error_str).

    Intenta primero con body libre. Si Twilio rechaza con error ContentSid,
    reintenta automáticamente con TWILIO_CONTENT_SID (o el content_sid
    que se pase explícitamente como override).
    """
    client = _get_twilio_client()
    if not client:
        return None, "Twilio no está configurado."
    numero = re.sub(r"[^\d+]", "", celular)
    if not numero.startswith("+"):
        numero = "+" + numero
    to_wa = f"whatsapp:{numero}"

    # Si se pasa un SID explícito (override manual), usarlo directo.
    if content_sid:
        return _enviar_via_template(client, to_wa, mensaje, content_sid)

    # Intento 1: mensaje libre (body)
    try:
        msg = client.messages.create(from_=TWILIO_WA_FROM, to=to_wa, body=mensaje)
        return msg.sid, None
    except Exception as e:
        err = str(e)
        # Intento 2: si el error es por ContentSid, reintentar con la plantilla configurada
        if TWILIO_CONTENT_SID and _es_error_contentsid(err):
            print(f"[Twilio] body rechazado (ContentSid), reintentando con plantilla …", flush=True)
            return _enviar_via_template(client, to_wa, mensaje, TWILIO_CONTENT_SID)
        return None, err


def _twilio_error_legible(error_str):
    """Traduce errores comunes de Twilio a mensajes en español con instrucciones."""
    if not error_str:
        return error_str
    e = error_str.lower()
    if "contentsid" in e or "content_sid" in e or "63016" in e or " mbs " in e:
        return (
            "El sandbox requiere una plantilla aprobada (ContentSid). "
            "Solución rápida: enviá el mensaje de activación al número del sandbox "
            "desde el destinatario (+5491161394159 → escribe 'join [keyword]' al "
            "+17372508034) para abrir la ventana de 24 h. "
            "O ingresá un Content SID de plantilla en el panel (sección 'Plantilla Twilio')."
        )
    if "21211" in e or "not a valid whatsapp" in e:
        return "Número no válido para WhatsApp o no registrado en Twilio."
    if "21608" in e or "unverified" in e:
        return "Número no verificado en la cuenta trial. Verificalo en la consola de Twilio primero."
    if "21614" in e or "not a mobile number" in e:
        return "El número no parece ser un celular válido."
    return error_str


def _guardar_envio(db, batch_id, cliente_id, cliente_nombre, celular,
                   mensaje, twilio_sid, estado, error_msg, es_prueba):
    db.execute(
        """INSERT INTO envios
           (batch_id, cliente_id, cliente_nombre, celular, mensaje,
            twilio_sid, estado, error_msg, es_prueba)
           VALUES (?,?,?,?,?,?,?,?,?)""",
        (batch_id, cliente_id, cliente_nombre, celular, mensaje,
         twilio_sid, estado, error_msg, es_prueba),
    )


@app.route("/whatsapp/enviar-prueba", methods=["POST"])
@login_required
def whatsapp_enviar_prueba():
    if not _twilio_configurado():
        flash("Configurá las variables de Twilio antes de enviar.", "danger")
        return redirect(url_for("whatsapp_preview"))

    mensaje_tpl = request.form.get("mensaje", "").strip()
    if not mensaje_tpl:
        flash("El mensaje está vacío.", "danger")
        return redirect(url_for("whatsapp_preview"))

    if not TWILIO_WA_OWNER:
        flash("No hay número de prueba configurado (TWILIO_WA_OWNER).", "danger")
        return redirect(url_for("whatsapp_preview"))

    content_sid = request.form.get("content_sid", "").strip() or None
    mensaje = mensaje_tpl.replace("{nombre}", "Prueba")
    sid, error = _enviar_whatsapp(TWILIO_WA_OWNER, mensaje, content_sid=content_sid)
    error_legible = _twilio_error_legible(error)

    db = get_db()
    _guardar_envio(db, None, None, "PRUEBA", TWILIO_WA_OWNER,
                   mensaje, sid, "enviado" if sid else "fallido", error_legible, 1)
    db.commit()
    db.close()

    if error_legible:
        flash(error_legible, "danger")
    else:
        flash(f"Mensaje de prueba enviado a {TWILIO_WA_OWNER}. ✅", "success")
    return redirect(url_for("whatsapp_preview"))


@app.route("/whatsapp/enviar-todos", methods=["POST"])
@login_required
def whatsapp_enviar_todos():
    if not _twilio_configurado():
        flash("Configurá las variables de Twilio antes de enviar.", "danger")
        return redirect(url_for("whatsapp_preview"))

    mensaje_tpl = request.form.get("mensaje", "").strip()
    if not mensaje_tpl:
        flash("El mensaje está vacío.", "danger")
        return redirect(url_for("whatsapp_preview"))

    db = get_db()
    clientes = [dict(r) for r in
                db.execute("SELECT * FROM clientes WHERE activo=1 ORDER BY nombre").fetchall()]

    if not clientes:
        db.close()
        flash("No hay clientes activos para enviar.", "warning")
        return redirect(url_for("whatsapp_preview"))

    content_sid = request.form.get("content_sid", "").strip() or None
    batch_id = uuid.uuid4().hex
    enviados = fallidos = 0

    for c in clientes:
        mensaje = mensaje_tpl.replace("{nombre}", c["nombre"])
        sid, error = _enviar_whatsapp(c["celular"], mensaje, content_sid=content_sid)
        estado = "enviado" if sid else "fallido"
        _guardar_envio(db, batch_id, c["id"], c["nombre"], c["celular"],
                       mensaje, sid, estado, _twilio_error_legible(error), 0)
        if sid:
            enviados += 1
        else:
            fallidos += 1
        time.sleep(0.3)

    db.commit()
    db.close()

    if fallidos == 0:
        flash(f"✅ Mensajes enviados a {enviados} cliente(s).", "success")
    else:
        flash(f"Enviados: {enviados} | Fallidos: {fallidos}", "warning")
    return redirect(url_for("whatsapp_historial"))


@app.route("/whatsapp/plantillas")
@login_required
def whatsapp_plantillas():
    """Devuelve JSON con las plantillas de contenido disponibles en la cuenta Twilio."""
    client = _get_twilio_client()
    if not client:
        return {"error": "Twilio no configurado"}, 400
    try:
        contents = client.content.v1.contents.list(limit=50)
        plantillas = []
        for c in contents:
            tipos = list(c.types.keys()) if c.types else []
            plantillas.append({
                "sid": c.sid,
                "nombre": c.friendly_name,
                "tipos": tipos,
            })
        return {"plantillas": plantillas}
    except Exception as e:
        return {"error": str(e)}, 500


@app.route("/whatsapp/historial")
@login_required
def whatsapp_historial():
    db = get_db()
    envios = [dict(r) for r in db.execute(
        "SELECT * FROM envios ORDER BY created_at DESC LIMIT 300"
    ).fetchall()]
    db.close()
    return render_template("whatsapp_historial.html", envios=envios)


# ── Importación masiva ────────────────────────────────────────────────────────

IMPORT_CACHE_DIR = "/tmp/mis_ventas_import"
VERDADERO = {"si", "sí", "yes", "1", "true", "verdadero", "v", "s", "y"}


def _parsear_en_oferta(val):
    return 1 if str(val).strip().lower() in VERDADERO else 0


def _parsear_unidad(val):
    return _UNIDAD_MAP.get(str(val or "").strip().lower(), "por unidad")


def _parsear_precio(val):
    try:
        return float(str(val).strip().replace(",", ".").replace("$", "").replace(" ", ""))
    except (ValueError, TypeError):
        return None


def _filas_excel(file_bytes):
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    return rows


def _filas_csv(file_bytes):
    text = file_bytes.decode("utf-8-sig")
    reader = csv.reader(io.StringIO(text))
    return list(reader)


def _fingerprint(headers):
    key = "|".join(sorted(h.strip().lower() for h in headers if h))
    return hashlib.sha256(key.encode()).hexdigest()[:16]


def _auto_detectar(headers):
    PATRONES = {
        "nombre":      ("nombre", "producto", "name", "artículo", "articulo", "item"),
        "descripcion": ("descrip", "detalle", "detail", "observ", "nota", "comment"),
        "precio":      ("precio", "price", "valor", "costo", "importe", "monto", "pvp", "venta"),
        "unidad":      ("unidad", "unit", "medida", "formato", "presentac"),
        "en_oferta":   ("oferta", "offer", "descuento", "promo", "rebaj", "sale"),
    }
    lower_h = [(i, h, h.strip().lower()) for i, h in enumerate(headers) if h]
    mapping, usados = {}, set()
    for campo, patrones in PATRONES.items():
        for i, h, lh in lower_h:
            if i not in usados and any(p in lh for p in patrones):
                mapping[campo] = h
                usados.add(i)
                break
    return mapping


def _get_saved_mapping(fingerprint):
    db = get_db()
    row = db.execute(
        "SELECT mapping FROM column_mappings WHERE fingerprint=?", (fingerprint,)
    ).fetchone()
    db.close()
    return json.loads(row["mapping"]) if row else None


def _save_mapping_db(fingerprint, mapping):
    db = get_db()
    db.execute(
        """INSERT INTO column_mappings (fingerprint, mapping) VALUES (?, ?)
           ON CONFLICT(fingerprint) DO UPDATE SET
             mapping=excluded.mapping,
             updated_at=datetime('now','localtime')""",
        (fingerprint, json.dumps(mapping, ensure_ascii=False)),
    )
    db.commit()
    db.close()


def _procesar_filas_con_mapping(data_rows, headers, field_to_col):
    header_idx = {h: i for i, h in enumerate(headers)}

    def get_val(row, campo):
        col = field_to_col.get(campo, "")
        if not col or col not in header_idx:
            return None
        idx = header_idx[col]
        return row[idx] if idx < len(row) else None

    def get_by_name(row, colname):
        if colname not in header_idx:
            return None
        idx = header_idx[colname]
        return row[idx] if idx < len(row) else None

    if not field_to_col.get("nombre") or not field_to_col.get("precio"):
        return [], ["El mapeo debe incluir los campos obligatorios: Nombre y Precio."]

    productos, errores = [], []
    for n, row in enumerate(data_rows, start=2):
        if all(str(c).strip() == "" for c in row):
            continue

        nombre = str(get_val(row, "nombre") or "").strip()
        if not nombre:
            errores.append(f"Fila {n}: nombre vacío — se omitió.")
            continue

        precio_raw = get_val(row, "precio")
        precio = _parsear_precio(precio_raw)
        if precio is None:
            errores.append(f"Fila {n} ({nombre!r}): precio inválido ({precio_raw!s}) — se omitió.")
            continue

        # Variantes adicionales (columnas unidad2/precio2, unidad3/precio3, etc.)
        extra_variantes = []
        for i in range(2, 5):
            precio_extra = _parsear_precio(get_by_name(row, f"precio{i}"))
            if precio_extra is not None:
                unidad_extra = _parsear_unidad(get_by_name(row, f"unidad{i}") or "")
                extra_variantes.append({"unidad": unidad_extra, "precio": precio_extra, "en_oferta": 0})

        productos.append({
            "nombre": nombre,
            "descripcion": str(get_val(row, "descripcion") or "").strip(),
            "precio": precio,
            "en_oferta": _parsear_en_oferta(get_val(row, "en_oferta") or ""),
            "unidad": _parsear_unidad(get_val(row, "unidad") or ""),
            "extra_variantes": extra_variantes,
        })

    return productos, errores


def _guardar_importados(productos):
    db = get_db()
    ids = []
    for p in productos:
        cur = db.execute(
            """INSERT INTO productos (nombre, descripcion, precio, unidad, en_oferta, activo)
               VALUES (?,?,?,?,?,1)""",
            (p["nombre"], p["descripcion"], p["precio"], p["unidad"], p["en_oferta"]),
        )
        pid = cur.lastrowid
        ids.append(pid)
        variantes = [{"unidad": p["unidad"], "precio": p["precio"], "en_oferta": p["en_oferta"], "orden": 0}]
        for i, ev in enumerate(p.get("extra_variantes", []), start=1):
            variantes.append({"unidad": ev["unidad"], "precio": ev["precio"], "en_oferta": 0, "orden": i})
        _guardar_variantes_db(pid, variantes, db)
    db.commit()
    placeholders = ",".join("?" * len(ids))
    nuevos = [dict(r) for r in db.execute(
        f"SELECT * FROM productos WHERE id IN ({placeholders}) ORDER BY nombre", ids
    ).fetchall()]
    db.close()
    return nuevos


@app.route("/productos/importar", methods=["GET", "POST"])
@login_required
def productos_importar():
    if request.method == "GET":
        return render_template("importar_productos.html")

    file = request.files.get("archivo")
    if not file or not file.filename:
        flash("Seleccioná un archivo.", "danger")
        return render_template("importar_productos.html")

    ext = file.filename.rsplit(".", 1)[-1].lower()
    if ext not in ("xlsx", "xls", "csv"):
        flash("Formato no soportado. Usá .xlsx o .csv.", "danger")
        return render_template("importar_productos.html")

    file_bytes = file.read()
    try:
        all_rows = _filas_excel(file_bytes) if ext in ("xlsx", "xls") else _filas_csv(file_bytes)
    except Exception as e:
        flash(f"No se pudo leer el archivo: {e}", "danger")
        return render_template("importar_productos.html")

    if len(all_rows) < 2:
        flash("El archivo necesita al menos una fila de encabezado y una de datos.", "warning")
        return render_template("importar_productos.html")

    headers = [str(c).strip() if c is not None else "" for c in all_rows[0]]
    data_rows = [[str(c) if c is not None else "" for c in row] for row in all_rows[1:]]

    os.makedirs(IMPORT_CACHE_DIR, exist_ok=True)
    tmp_key = uuid.uuid4().hex
    with open(os.path.join(IMPORT_CACHE_DIR, f"{tmp_key}.json"), "w", encoding="utf-8") as f:
        json.dump({"headers": headers, "rows": data_rows}, f, ensure_ascii=False)

    session["import_state"] = {
        "tmp_key": tmp_key,
        "fingerprint": _fingerprint(headers),
        "filename": file.filename,
    }
    return redirect(url_for("productos_mapear"))


@app.route("/productos/importar/mapear", methods=["GET", "POST"])
@login_required
def productos_mapear():
    state = session.get("import_state")
    if not state:
        flash("La sesión de importación expiró. Subí el archivo nuevamente.", "warning")
        return redirect(url_for("productos_importar"))

    tmp_path = os.path.join(IMPORT_CACHE_DIR, f"{state['tmp_key']}.json")
    try:
        with open(tmp_path, encoding="utf-8") as f:
            cached = json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        flash("El archivo temporal no se encontró. Subí el archivo nuevamente.", "danger")
        session.pop("import_state", None)
        return redirect(url_for("productos_importar"))

    headers = cached["headers"]
    fingerprint = state["fingerprint"]

    if request.method == "GET":
        saved = _get_saved_mapping(fingerprint)
        if saved:
            saved = {k: v for k, v in saved.items() if not v or v in headers}
        mapping = saved or _auto_detectar(headers)
        preview = cached["rows"][:5]
        return render_template(
            "mapear_columnas.html",
            state=state,
            headers=headers,
            preview=preview,
            mapping=mapping,
            tiene_guardado=bool(saved),
        )

    CAMPOS = ("nombre", "descripcion", "precio", "unidad", "en_oferta")
    field_to_col = {
        campo: request.form.get(f"map_{campo}", "").strip()
        for campo in CAMPOS
        if request.form.get(f"map_{campo}", "").strip()
    }

    if not field_to_col.get("nombre") or not field_to_col.get("precio"):
        flash("Debés mapear los campos obligatorios: Nombre y Precio.", "danger")
        preview = cached["rows"][:5]
        return render_template(
            "mapear_columnas.html",
            state=state,
            headers=headers,
            preview=preview,
            mapping=field_to_col,
            tiene_guardado=False,
        )

    if request.form.get("guardar_mapeo"):
        _save_mapping_db(fingerprint, field_to_col)

    productos, errores = _procesar_filas_con_mapping(cached["rows"], headers, field_to_col)

    try:
        os.remove(tmp_path)
    except OSError:
        pass
    session.pop("import_state", None)

    if not productos:
        for err in errores:
            flash(err, "danger")
        return redirect(url_for("productos_importar"))

    nuevos = _guardar_importados(productos)
    for err in errores:
        flash(err, "warning")
    flash(f"Se importaron {len(productos)} productos correctamente.", "success")
    return render_template("importar_resultado.html", productos=nuevos, errores=errores)


@app.route("/productos/plantilla")
@login_required
def productos_plantilla():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Productos"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="0D6EFD")
    center = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    labels = [
        "Nombre *", "Descripción", "Precio 1 *", "En oferta (sí/no)",
        "Unidad 1", "Unidad 2", "Precio 2", "Unidad 3", "Precio 3",
    ]
    widths = [30, 35, 12, 18, 16, 14, 12, 14, 12]

    for col, (label, width) in enumerate(zip(labels, widths), start=1):
        cell = ws.cell(row=1, column=col, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border
        ws.column_dimensions[cell.column_letter].width = width

    ws.row_dimensions[1].height = 22

    nota_fill = PatternFill("solid", fgColor="FFF9E6")
    ejemplos = [
        ("Empanadas de carne", "Docena, rellenas con cebolla y huevo",
         2500, "no", "por unidad", None, None, None, None),
        ("Queso cremoso", "Cremoso suave",
         3500, "sí", "por kg", "por media horma", 8000, "por horma entera", 15000),
        ("Salame", "Estacionado, picante",
         4200, "no", "por 100g", None, None, None, None),
        ("Medialunas", "Con manteca, crocantes",
         1200, "sí", "por unidad", None, None, None, None),
        ("Queso en horma", "Horma entera aprox. 3 kg",
         11000, "no", "por horma entera", None, None, None, None),
        ("Queso 1/4", "Un cuarto de horma",
         2800, "sí", "por 1/4 kg", None, None, None, None),
    ]
    for row_num, fila in enumerate(ejemplos, start=2):
        for col, val in enumerate(fila, start=1):
            if val is not None:
                cell = ws.cell(row=row_num, column=col, value=val)
                cell.fill = nota_fill
                cell.border = border

    ws2 = wb.create_sheet("Instrucciones")
    unidades_str = " / ".join(UNIDADES)
    instrucciones = [
        ("INSTRUCCIONES PARA COMPLETAR LA PLANTILLA", True),
        ("", False),
        ("Columna NOMBRE (obligatorio):", True),
        ("  Escribí el nombre del producto. No puede estar vacío.", False),
        ("", False),
        ("Columna DESCRIPCIÓN:", True),
        ("  Descripción corta (opcional). Puede quedar en blanco.", False),
        ("", False),
        ("Columna PRECIO 1 (obligatorio):", True),
        ("  Número con punto o coma decimal. Ej: 1500 o 1500,50", False),
        ("", False),
        ("Columna EN OFERTA:", True),
        ('  Escribí "sí" o "no". También acepta: si / yes / 1 / no / 0.', False),
        ("", False),
        ("Columna UNIDAD 1 (y UNIDAD 2, UNIDAD 3):", True),
        (f"  Opciones válidas: {unidades_str}", False),
        ("  También se aceptan abreviaturas: kg, 100g, 1/4, horma, media horma, etc.", False),
        ("  Si dejás vacío, se usa 'por unidad'.", False),
        ("", False),
        ("VARIANTES DE PRECIO:", True),
        ("  Para agregar un segundo precio, completá UNIDAD 2 y PRECIO 2.", False),
        ("  Para un tercero, completá UNIDAD 3 y PRECIO 3.", False),
        ("  Ejemplo: Queso cremoso → $3500/kg, $8000/media horma, $15000/horma entera", False),
        ("", False),
        ("IMPORTANTE:", True),
        ("  - La primera fila DEBE ser el encabezado (no la borres).", False),
        ("  - Las fotos las agregás desde el panel, una por una, después de importar.", False),
        ("  - Podés importar el archivo como .xlsx o guardarlo como .csv.", False),
    ]
    ws2.column_dimensions["A"].width = 70
    for i, (texto, bold) in enumerate(instrucciones, start=1):
        cell = ws2.cell(row=i, column=1, value=texto)
        cell.font = Font(bold=bold, size=11 if bold else 10)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return Response(
        output.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=plantilla_productos.xlsx"},
    )


# ── Archivos estáticos subidos ─────────────────────────────────────────────────

@app.route("/uploads/<filename>")
def uploaded_file(filename):
    return send_from_directory(UPLOAD_FOLDER, filename)


if __name__ == "__main__":
    init_db()
    os.makedirs(UPLOAD_FOLDER, exist_ok=True)
    os.makedirs(IMPORT_CACHE_DIR, exist_ok=True)
    app.run(debug=True, host="0.0.0.0", port=5000)
